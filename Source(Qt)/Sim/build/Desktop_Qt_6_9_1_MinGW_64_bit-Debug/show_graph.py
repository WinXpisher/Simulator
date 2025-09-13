import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
from collections import defaultdict

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import Font

# ---------------------------
# Функція аналізу й експорту одного логу
# ---------------------------
def parse_and_export(log_path: str, method_name: str) -> pd.DataFrame:
    run_re  = re.compile(
        r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\s+\[RUN\] Task \[(.*?)\] started running on resource (\w+)'
    )
    perf_re = re.compile(
        r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\s+\[PERFORMED\] Task \[(.*?)\] has been performed'
    )

    start_times     = {}
    task_resource   = {}
    resource_counts = defaultdict(int)
    resource_times  = defaultdict(float)
    first_ts = last_ts = None

    with open(log_path, encoding="utf-8") as f:
        for line in f:
            if m := run_re.search(line):
                ts_str, tid, res = m.groups()
                ts = datetime.strptime(ts_str, '%Y-%m-%d %H:%M:%S')
                first_ts = ts if first_ts is None or ts < first_ts else first_ts
                last_ts  = ts if last_ts  is None or ts > last_ts  else last_ts
                start_times[tid]   = ts
                task_resource[tid] = res
            elif m := perf_re.search(line):
                ts_str, tid = m.groups()
                ts = datetime.strptime(ts_str, '%Y-%m-%d %H:%M:%S')
                first_ts = ts if first_ts is None or ts < first_ts else first_ts
                last_ts  = ts if last_ts  is None or ts > last_ts  else last_ts
                if tid in start_times:
                    duration = (ts - start_times.pop(tid)).total_seconds()
                    res = task_resource.pop(tid)
                    resource_counts[res] += 1
                    resource_times[res]  += duration

    Tk_total = (last_ts - first_ts).total_seconds() if first_ts and last_ts else 0
    resources = sorted(resource_counts)
    K_i       = [resource_counts[r] for r in resources]
    tk_i      = [resource_times[r]  for r in resources]
    RT_i      = [(t / Tk_total * 100) if Tk_total else 0 for t in tk_i]
    DT_i      = [abs((Tk_total - t) / Tk_total * 100) if Tk_total else 0 for t in tk_i]

    df = pd.DataFrame({
        "Resource":    resources,
        "K_i":         K_i,
        "tk_i (sec)":  tk_i,
        "RT_i (%)":    RT_i,
        "DT_i (%)":    DT_i
    })

    total = pd.DataFrame({
        "Resource":    ["Total"],
        "K_i":         [sum(K_i)],
        "tk_i (sec)":  [sum(tk_i)],
        "RT_i (%)":    [sum(tk_i)/Tk_total/len(resources)*100 if Tk_total and resources else 0],
        "DT_i (%)":    [abs((Tk_total - sum(tk_i))/Tk_total*100) if Tk_total else 0]
    })
    df = pd.concat([df, total], ignore_index=True)

    base       = os.path.splitext(os.path.basename(log_path))[0]
    log_number = base.split('_')[-1]
    out_dir    = "Graphs"
    os.makedirs(out_dir, exist_ok=True)
    excel_path = os.path.join(out_dir, f"{method_name}_{log_number}.xlsx")

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Summary")
        ws   = writer.sheets["Summary"]
        bold = Font(bold=True)
        for col in range(1, df.shape[1] + 1):
            ws.cell(row=1,               column=col).font = bold
            ws.cell(row=df.shape[0] + 1, column=col).font = bold
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = max_len + 2

    return df

# ---------------------------
# Функція відкриття логу з урахуванням вибраного методу
# ---------------------------
def open_log():
    METHOD_FOLDERS = {
        "LIFO":     "lifo",
        "FIFO":     "fcfs",
        "HPF":      "hpf",
        "SMART":    "smart",
        "PENGUIN":  "penguin",
        "MFQS":     "mfqs",
        "SIMPLEX":  "simplex",
        "BACKFILL": "backfill"
    }

    # Беремо вибір безпосередньо з Combobox
    method = combobox.get().strip()
    if method not in METHOD_FOLDERS:
        messagebox.showerror("Помилка", "Будь ласка, оберіть метод")
        return

    folder_name = METHOD_FOLDERS[method]
    folder = os.path.join(os.getcwd(), folder_name)
    if not os.path.isdir(folder):
        messagebox.showerror("Помилка", f"Папка '{folder_name}' не існує")
        return

    path = filedialog.askopenfilename(
        title=f"Виберіть лог ({method})",
        filetypes=[("Text files", "*.txt")],
        initialdir=folder
    )
    if not path:
        return

    # Аналіз та експорт
    parse_and_export(path, method)

    # Збір даних для гістограми
    clock_re = re.compile(r"\[!\]\s*Clock:\s*(\d+)")
    run_re_simple = re.compile(r"\[RUN\] Task \[(.*?)\] started")
    graph_data    = defaultdict(lambda: {"exec": 0, "queue": 0})
    current_clock = None

    with open(path, encoding="utf-8") as f:
        for line in f:
            if m := clock_re.search(line):
                current_clock = int(m.group(1))
            elif current_clock is not None:
                if run_re_simple.search(line):
                    h = current_clock // 60 + 1
                    graph_data[h]["exec"] += 1
                elif "[SEND]" in line:
                    h = current_clock // 60 + 1
                    graph_data[h]["queue"] += 1

    hours        = sorted(graph_data)
    exec_counts  = [graph_data[h]["exec"]  for h in hours]
    queue_counts = [graph_data[h]["queue"] for h in hours]

    # Побудова бар-чарту
    x = range(len(hours))
    w = 0.4
    plt.figure(figsize=(8, 6))
    plt.bar([i - w/2 for i in x], exec_counts,  width=w, label="Виконуються", color="C0")
    plt.bar([i + w/2 for i in x], queue_counts, width=w, label="У черзі",       color="orange")
    base = os.path.splitext(os.path.basename(path))[0]
    log_number = base.split('_')[-1]
    plt.title(f"Гістограма (Лог: {log_number}, Метод: {method})")
    plt.xlabel("Години")
    plt.ylabel("Кількість задач")
    plt.xticks(x, hours)
    plt.legend()
    plt.grid(axis="y", linestyle="--", alpha=0.5)
    plt.tight_layout()
    plt.show()

    # Збереження даних для графіка
    df_g = pd.DataFrame({
        "Години":      hours,
        "Виконуються": exec_counts,
        "В черзі":     queue_counts
    })
    graph_path = os.path.join("Graphs", f"graph_info_{log_number}_{method}.xlsx")
    with pd.ExcelWriter(graph_path, engine="openpyxl") as writer:
        df_g.to_excel(writer, index=False, sheet_name="Graph Info")
        ws   = writer.sheets["Graph Info"]
        bold = Font(bold=True)
        for col in range(1, df_g.shape[1] + 1):
            ws.cell(row=1, column=col).font = bold
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = max_len + 2

# ---------------------------
# Налаштування GUI
# ---------------------------
root = tk.Tk()
root.title("Обробка логу")
root.resizable(False, False)

# Центруємо вікно
W, H = 500, 200
sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")

methods = ["LIFO", "FIFO", "HPF", "SMART", "PENGUIN", "MFQS", "SIMPLEX", "BACKFILL"]

# StringVar без жорсткого дефолту
method_var = tk.StringVar()

ttk.Style().theme_use('clam')
ttk.Label(root, text="Оберіть метод:").pack(anchor="w", padx=10, pady=(20, 5))

combobox = ttk.Combobox(
    root,
    textvariable=method_var,
    values=methods,
    state="readonly"
)
combobox.pack(fill="x", padx=10)

# Встановлюємо початковий індекс (0 => "LIFO")
combobox.current(0)

ttk.Button(
    root,
    text="Показати",
    command=open_log
).pack(pady=20, ipadx=10, ipady=5)

root.mainloop()